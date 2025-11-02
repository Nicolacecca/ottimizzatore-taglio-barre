import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import List, Tuple, Dict
import copy
import random
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class OttimizzatoreTaglio:
    """Algoritmo per ottimizzare il taglio di barre minimizzando gli scarti"""

    def __init__(self, barre_disponibili: List[Tuple[int, float]], spessore_lama: float):
        """
        Args:
            barre_disponibili: Lista di tuple (quantità, lunghezza) delle barre disponibili
            spessore_lama: Spessore della lama in mm
        """
        self.barre_disponibili = sorted(barre_disponibili, key=lambda x: x[1], reverse=True)
        self.spessore_lama = spessore_lama

    def ottimizza(self, pezzi_richiesti: List[Tuple[int, float]]) -> List[Dict]:
        """
        Ottimizza i tagli usando algoritmo First Fit Decreasing con supporto per barre di lunghezze diverse
        Ad ogni chiamata genera un pattern diverso introducendo variabilità nell'ordine dei pezzi

        Args:
            pezzi_richiesti: Lista di tuple (quantità, lunghezza)

        Returns:
            Lista di barre con i tagli ottimizzati
        """
        # Espandi i pezzi in base alla quantità
        tutti_pezzi = []
        for qty, lunghezza in pezzi_richiesti:
            tutti_pezzi.extend([lunghezza] * qty)

        # Ordina per lunghezza decrescente con piccola variazione casuale
        # Raggruppa pezzi per lunghezza e mescola all'interno dei gruppi
        pezzi_per_lunghezza = {}
        for pezzo in tutti_pezzi:
            if pezzo not in pezzi_per_lunghezza:
                pezzi_per_lunghezza[pezzo] = []
            pezzi_per_lunghezza[pezzo].append(pezzo)

        # Ricostruisci la lista con ordine variabile
        tutti_pezzi = []
        lunghezze_ordinate = sorted(pezzi_per_lunghezza.keys(), reverse=True)

        # Mescola le lunghezze simili per creare variabilità
        i = 0
        while i < len(lunghezze_ordinate):
            gruppo = [lunghezze_ordinate[i]]
            j = i + 1
            # Raggruppa lunghezze entro il 10% di differenza
            while j < len(lunghezze_ordinate) and lunghezze_ordinate[j] >= lunghezze_ordinate[i] * 0.9:
                gruppo.append(lunghezze_ordinate[j])
                j += 1

            # Mescola il gruppo
            random.shuffle(gruppo)
            for lung in gruppo:
                tutti_pezzi.extend(pezzi_per_lunghezza[lung])

            i = j

        # Crea pool di barre disponibili
        barre_pool = []
        for qty, lunghezza in self.barre_disponibili:
            for _ in range(qty):
                barre_pool.append(lunghezza)

        barre_utilizzate = []

        for pezzo in tutti_pezzi:
            # Cerca una barra già in uso dove il pezzo può essere inserito
            inserito = False

            # Trova tutte le barre compatibili
            barre_compatibili = [(idx, barra) for idx, barra in enumerate(barre_utilizzate)
                                 if pezzo <= barra['spazio_rimanente']]

            if barre_compatibili:
                # Scegli una barra casualmente tra quelle con spazio sufficiente
                # Questo crea pattern diversi ad ogni ottimizzazione
                idx_scelto, barra_scelta = random.choice(barre_compatibili)
                barra_scelta['tagli'].append(pezzo)
                barra_scelta['spazio_rimanente'] -= (pezzo + self.spessore_lama)
                barra_scelta['num_tagli'] += 1
                inserito = True

            # Se non è stato inserito, prendi una nuova barra dal pool
            if not inserito:
                # Trova la barra più piccola che può contenere il pezzo
                # Questo risparmia le barre più lunghe per pezzi più grandi
                barra_scelta = None
                barre_compatibili = [b for b in barre_pool if pezzo <= b]

                if not barre_compatibili:
                    if not barre_pool:
                        raise ValueError(
                            f"Barre disponibili esaurite!\n\n"
                            f"Servono più barre per completare tutti i tagli.\n"
                            f"Pezzo da inserire: {pezzo}mm\n"
                            f"Barre rimaste: 0\n\n"
                            f"Soluzione: Aggiungi più barre disponibili"
                        )
                    else:
                        raise ValueError(
                            f"Nessuna barra disponibile può contenere il pezzo da {pezzo}mm!\n\n"
                            f"Barra più lunga disponibile: {max(barre_pool)}mm\n"
                            f"Pezzo richiesto: {pezzo}mm\n\n"
                            f"Soluzione: Aggiungi barre più lunghe di almeno {pezzo}mm"
                        )

                # Scegli la barra più piccola tra quelle compatibili
                barra_scelta = min(barre_compatibili)

                # Rimuovi la barra dal pool
                barre_pool.remove(barra_scelta)

                nuova_barra = {
                    'lunghezza': barra_scelta,
                    'tagli': [pezzo],
                    'spazio_rimanente': barra_scelta - pezzo - self.spessore_lama,
                    'num_tagli': 1,
                    'sfrido': 0
                }
                barre_utilizzate.append(nuova_barra)

        # Calcola lo sfrido per ogni barra
        # Lo sfrido è lo spazio rimanente dopo l'ultimo pezzo
        # Non si aggiunge lo spessore lama perché dopo l'ultimo pezzo non si taglia più
        for barra in barre_utilizzate:
            barra['sfrido'] = barra['spazio_rimanente']

        return barre_utilizzate


class ApplicativoGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Ottimizzatore Taglio Barre")
        self.root.geometry("1000x800")

        # Imposta icona personalizzata se disponibile
        self.imposta_icona()

        # Variabili
        self.pezzi_richiesti = []
        self.barre_disponibili = []
        self.risultati_ottimizzazione = []
        self.modalita = tk.StringVar(value="disponibili")  # "disponibili" o "calcola"
        self.lunghezze_catalogo = []  # Solo lunghezze per modalità calcola
        self.costi_barre = {}  # Dict {lunghezza: costo} opzionale
        self.costo_barre_intere = 0  # Costo totale barre intere
        self.costo_effettivo = 0  # Costo effettivo basato su lunghezza utilizzata

        # Frame principale
        self.setup_ui()

    def imposta_icona(self, finestra=None):
        """Imposta l'icona della finestra

        Args:
            finestra: Finestra su cui impostare l'icona. Se None, usa self.root
        """
        if finestra is None:
            finestra = self.root

        try:
            # Cerca file icona nella stessa cartella dello script
            script_dir = os.path.dirname(os.path.abspath(__file__))
            ico_path = os.path.join(script_dir, "icon.ico")
            png_path = os.path.join(script_dir, "icon.png")

            if os.path.exists(ico_path):
                # Usa file .ico se esiste
                finestra.iconbitmap(ico_path)
            elif os.path.exists(png_path):
                # Usa file .png se esiste
                icon = tk.PhotoImage(file=png_path)
                finestra.iconphoto(True, icon)
            # Se non trova nessun file, mantiene l'icona di default
        except Exception:
            # Se c'è un errore, ignora e usa l'icona di default
            pass

    def setup_ui(self):
        # Frame parametri
        frame_params = ttk.LabelFrame(self.root, text="Parametri", padding=10)
        frame_params.pack(fill="x", padx=10, pady=5)

        ttk.Label(frame_params, text="Spessore lama (mm):").grid(row=0, column=0, sticky="w", padx=5)
        self.entry_spessore_lama = ttk.Entry(frame_params, width=15)
        self.entry_spessore_lama.insert(0, "3")
        self.entry_spessore_lama.grid(row=0, column=1, padx=5)

        ttk.Button(frame_params, text="?", command=self.mostra_help, width=3).grid(row=0, column=2, padx=20)

        # Pulsante per creare file Excel di esempio
        ttk.Button(frame_params, text="Crea Excel Esempio", command=self.crea_excel_esempio, width=20).grid(row=0, column=7, padx=5)

        # Selezione modalità
        ttk.Label(frame_params, text="Modalità:").grid(row=0, column=3, sticky="w", padx=(30, 5))
        ttk.Radiobutton(frame_params, text="Barre disponibili", variable=self.modalita,
                       value="disponibili", command=self.cambia_modalita).grid(row=0, column=4, padx=5)
        ttk.Radiobutton(frame_params, text="Calcola fabbisogno", variable=self.modalita,
                       value="calcola", command=self.cambia_modalita).grid(row=0, column=5, padx=5)

        # Frame container per le due sezioni affiancate
        frame_input_container = ttk.Frame(self.root)
        frame_input_container.pack(fill="both", expand=True, padx=10, pady=5)

        # ===== COLONNA SINISTRA: Barre Disponibili =====
        frame_barre_col = ttk.Frame(frame_input_container)
        frame_barre_col.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # Frame barre disponibili
        self.frame_barre = ttk.LabelFrame(frame_barre_col, text="Barre Disponibili", padding=10)
        self.frame_barre.pack(fill="x")

        self.label_qty_barra = ttk.Label(self.frame_barre, text="Quantità:")
        self.label_qty_barra.grid(row=0, column=0, sticky="w", padx=5)
        self.entry_qty_barra = ttk.Entry(self.frame_barre, width=8)
        self.entry_qty_barra.grid(row=0, column=1, padx=5)

        ttk.Label(self.frame_barre, text="Lunghezza (mm):").grid(row=0, column=2, sticky="w", padx=5)
        self.entry_lung_barra = ttk.Entry(self.frame_barre, width=12)
        self.entry_lung_barra.grid(row=0, column=3, padx=5)

        ttk.Label(self.frame_barre, text="Costo (€):").grid(row=0, column=4, sticky="w", padx=5)
        self.entry_costo_barra = ttk.Entry(self.frame_barre, width=10)
        self.entry_costo_barra.grid(row=0, column=5, padx=5)
        ttk.Label(self.frame_barre, text="(opzionale)", font=("Arial", 7, "italic")).grid(row=0, column=6, sticky="w", padx=2)

        ttk.Button(self.frame_barre, text="Aggiungi", command=self.aggiungi_barra, width=10).grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="ew")
        ttk.Button(self.frame_barre, text="Rimuovi", command=self.rimuovi_barra, width=10).grid(row=1, column=2, columnspan=2, padx=5, pady=5, sticky="ew")
        ttk.Button(self.frame_barre, text="Importa Excel", command=self.importa_barre_excel, width=12).grid(row=1, column=4, columnspan=3, padx=5, pady=5, sticky="ew")

        # Lista barre
        frame_lista_barre = ttk.Frame(frame_barre_col)
        frame_lista_barre.pack(fill="both", expand=True, pady=(5, 0))

        self.tree_barre = ttk.Treeview(frame_lista_barre, columns=("Quantità", "Lunghezza", "Costo"), show="headings", height=5)
        self.tree_barre.heading("Quantità", text="Quantità")
        self.tree_barre.heading("Lunghezza", text="Lunghezza (mm)")
        self.tree_barre.heading("Costo", text="Costo (€)")
        self.tree_barre.column("Quantità", width=70)
        self.tree_barre.column("Lunghezza", width=100)
        self.tree_barre.column("Costo", width=80)

        scrollbar_barre = ttk.Scrollbar(frame_lista_barre, orient="vertical", command=self.tree_barre.yview)
        self.tree_barre.configure(yscrollcommand=scrollbar_barre.set)

        self.tree_barre.pack(side="left", fill="both", expand=True)
        scrollbar_barre.pack(side="right", fill="y")

        # ===== COLONNA DESTRA: Pezzi Richiesti =====
        frame_pezzi_col = ttk.Frame(frame_input_container)
        frame_pezzi_col.pack(side="left", fill="both", expand=True, padx=(5, 0))

        # Frame input pezzi
        frame_input = ttk.LabelFrame(frame_pezzi_col, text="Pezzi Richiesti", padding=10)
        frame_input.pack(fill="x")

        ttk.Label(frame_input, text="Quantità:").grid(row=0, column=0, sticky="w", padx=5)
        self.entry_quantita = ttk.Entry(frame_input, width=8)
        self.entry_quantita.grid(row=0, column=1, padx=5)

        ttk.Label(frame_input, text="Lunghezza (mm):").grid(row=0, column=2, sticky="w", padx=5)
        self.entry_lunghezza = ttk.Entry(frame_input, width=12)
        self.entry_lunghezza.grid(row=0, column=3, padx=5)

        # Pulsanti su un'unica riga
        ttk.Button(frame_input, text="Aggiungi", command=self.aggiungi_pezzo, width=8).grid(row=1, column=0, padx=(5,2), pady=5, sticky="ew")
        ttk.Button(frame_input, text="Rimuovi", command=self.rimuovi_pezzo, width=8).grid(row=1, column=1, padx=2, pady=5, sticky="ew")
        ttk.Button(frame_input, text="Pulisci", command=self.pulisci_pezzi, width=8).grid(row=1, column=2, padx=2, pady=5, sticky="ew")
        ttk.Button(frame_input, text="Importa Excel", command=self.importa_pezzi_excel, width=12).grid(row=1, column=3, padx=(2,5), pady=5, sticky="ew")

        # Lista pezzi
        frame_lista = ttk.Frame(frame_pezzi_col)
        frame_lista.pack(fill="both", expand=True, pady=(5, 0))

        self.tree_pezzi = ttk.Treeview(frame_lista, columns=("Quantità", "Lunghezza"), show="headings", height=5)
        self.tree_pezzi.heading("Quantità", text="Quantità")
        self.tree_pezzi.heading("Lunghezza", text="Lunghezza (mm)")
        self.tree_pezzi.column("Quantità", width=80)
        self.tree_pezzi.column("Lunghezza", width=120)

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=self.tree_pezzi.yview)
        self.tree_pezzi.configure(yscrollcommand=scrollbar.set)

        self.tree_pezzi.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame risultati
        frame_risultati = ttk.LabelFrame(self.root, text="Risultati", padding=10)
        frame_risultati.pack(fill="both", expand=True, padx=10, pady=5)

        # Frame superiore con pulsanti e statistiche
        frame_top_risultati = ttk.Frame(frame_risultati)
        frame_top_risultati.pack(fill="x", pady=(0, 5))

        # Pulsanti azione a sinistra
        frame_azioni = ttk.Frame(frame_top_risultati)
        frame_azioni.pack(side="left")

        ttk.Button(frame_azioni, text="OTTIMIZZA", command=self.ottimizza, width=18).pack(side="left", padx=(0, 5))
        ttk.Button(frame_azioni, text="GENERA PDF", command=self.genera_pdf, width=18).pack(side="left")

        # Statistiche a destra
        self.label_stats = ttk.Label(frame_top_risultati, text="", font=("Arial", 9, "bold"))
        self.label_stats.pack(side="left", padx=(20, 0))

        # Tabella risultati
        self.tree_risultati = ttk.Treeview(frame_risultati, columns=("Barra", "Lunghezza", "Tagli", "Num Tagli", "Sfrido"), show="headings", height=8)
        self.tree_risultati.heading("Barra", text="Barra #")
        self.tree_risultati.heading("Lunghezza", text="Lung. Barra")
        self.tree_risultati.heading("Tagli", text="Tagli (mm)")
        self.tree_risultati.heading("Num Tagli", text="N° Tagli")
        self.tree_risultati.heading("Sfrido", text="Sfrido (mm)")

        self.tree_risultati.column("Barra", width=70)
        self.tree_risultati.column("Lunghezza", width=100)
        self.tree_risultati.column("Tagli", width=400)
        self.tree_risultati.column("Num Tagli", width=80)
        self.tree_risultati.column("Sfrido", width=100)

        scrollbar_ris = ttk.Scrollbar(frame_risultati, orient="vertical", command=self.tree_risultati.yview)
        self.tree_risultati.configure(yscrollcommand=scrollbar_ris.set)

        self.tree_risultati.pack(side="left", fill="both", expand=True)
        scrollbar_ris.pack(side="right", fill="y")

    def cambia_modalita(self):
        """Cambia tra modalità 'disponibili' e 'calcola fabbisogno'"""
        modalita = self.modalita.get()

        # Pulisci liste
        self.tree_barre.delete(*self.tree_barre.get_children())
        self.barre_disponibili = []
        self.lunghezze_catalogo = []
        self.costi_barre = {}
        self.entry_qty_barra.delete(0, tk.END)
        self.entry_lung_barra.delete(0, tk.END)
        self.entry_costo_barra.delete(0, tk.END)

        if modalita == "disponibili":
            # Modalità barre disponibili: mostra campo quantità
            self.frame_barre.config(text="Barre Disponibili")
            self.label_qty_barra.grid()
            self.entry_qty_barra.grid()
            self.tree_barre.heading("Quantità", text="Quantità")
            self.tree_barre.column("Quantità", width=70)
        else:
            # Modalità calcola: nascondi campo quantità
            self.frame_barre.config(text="Lunghezze Catalogo Barre")
            self.label_qty_barra.grid_remove()
            self.entry_qty_barra.grid_remove()
            self.tree_barre.heading("Quantità", text="Tipo")
            self.tree_barre.column("Quantità", width=70)

    def aggiungi_barra(self):
        try:
            modalita = self.modalita.get()
            lunghezza = float(self.entry_lung_barra.get())

            if lunghezza <= 0:
                messagebox.showerror("Errore", "La lunghezza deve essere maggiore di zero")
                return

            # Leggi costo opzionale
            costo_str = self.entry_costo_barra.get().strip()
            costo = None
            if costo_str:
                costo = float(costo_str)
                if costo < 0:
                    messagebox.showerror("Errore", "Il costo non può essere negativo")
                    return
                self.costi_barre[lunghezza] = costo

            if modalita == "disponibili":
                # Modalità normale: richiedi quantità
                qty = int(self.entry_qty_barra.get())
                if qty <= 0:
                    messagebox.showerror("Errore", "La quantità deve essere maggiore di zero")
                    return

                self.barre_disponibili.append((qty, lunghezza))
                costo_display = f"{costo:.2f}" if costo is not None else "-"
                self.tree_barre.insert("", "end", values=(qty, lunghezza, costo_display))
            else:
                # Modalità calcola: solo lunghezza
                if lunghezza in self.lunghezze_catalogo:
                    messagebox.showwarning("Attenzione", "Lunghezza già presente")
                    return

                self.lunghezze_catalogo.append(lunghezza)
                costo_display = f"{costo:.2f}" if costo is not None else "-"
                self.tree_barre.insert("", "end", values=("Catalogo", lunghezza, costo_display))

            self.entry_qty_barra.delete(0, tk.END)
            self.entry_lung_barra.delete(0, tk.END)
            self.entry_costo_barra.delete(0, tk.END)

        except ValueError:
            messagebox.showerror("Errore", "Inserire valori numerici validi")

    def rimuovi_barra(self):
        selected = self.tree_barre.selection()
        if selected:
            idx = self.tree_barre.index(selected[0])
            self.tree_barre.delete(selected[0])

            modalita = self.modalita.get()
            if modalita == "disponibili":
                _, lunghezza = self.barre_disponibili.pop(idx)
                if lunghezza in self.costi_barre:
                    del self.costi_barre[lunghezza]
            else:
                lunghezza = self.lunghezze_catalogo.pop(idx)
                if lunghezza in self.costi_barre:
                    del self.costi_barre[lunghezza]

    def aggiungi_pezzo(self):
        try:
            qty = int(self.entry_quantita.get())
            lunghezza = float(self.entry_lunghezza.get())

            if qty <= 0 or lunghezza <= 0:
                messagebox.showerror("Errore", "Quantità e lunghezza devono essere maggiori di zero")
                return

            self.pezzi_richiesti.append((qty, lunghezza))
            self.tree_pezzi.insert("", "end", values=(qty, lunghezza))

            self.entry_quantita.delete(0, tk.END)
            self.entry_lunghezza.delete(0, tk.END)

        except ValueError:
            messagebox.showerror("Errore", "Inserire valori numerici validi")

    def rimuovi_pezzo(self):
        selected = self.tree_pezzi.selection()
        if selected:
            idx = self.tree_pezzi.index(selected[0])
            self.tree_pezzi.delete(selected[0])
            self.pezzi_richiesti.pop(idx)

    def pulisci_pezzi(self):
        self.tree_pezzi.delete(*self.tree_pezzi.get_children())
        self.pezzi_richiesti = []

    def importa_barre_excel(self):
        """Importa barre disponibili da file Excel (ottimizzato per file grandi)"""
        filename = filedialog.askopenfilename(
            title="Seleziona file Excel con barre disponibili",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not filename:
            return

        try:
            # read_only=True per ottimizzare lettura file grandi
            wb = load_workbook(filename, data_only=True, read_only=True)

            # Cerca il foglio giusto
            ws = None
            for nome_foglio in ["Barre", "Magazzino", "Barre Disponibili", "Disponibili"]:
                if nome_foglio in wb.sheetnames:
                    ws = wb[nome_foglio]
                    break

            if ws is None:
                ws = wb.active

            modalita = self.modalita.get()
            righe_importate = 0
            errori = []

            # Batch insert per migliorare performance GUI
            batch_data = []
            batch_size = 50

            # Salta l'intestazione e leggi i dati
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue

                try:
                    if modalita == "disponibili":
                        qty = row[0]
                        lunghezza = row[1]
                        costo = row[2] if len(row) > 2 else None

                        if qty is None or lunghezza is None:
                            continue

                        qty = int(float(qty))
                        lunghezza = float(lunghezza)

                        if qty <= 0 or lunghezza <= 0:
                            if len(errori) < 100:  # Limita errori memorizzati
                                errori.append(f"Riga {idx}: valori devono essere positivi")
                            continue

                        if costo is not None and costo != "":
                            costo = float(costo)
                            if costo >= 0:
                                self.costi_barre[lunghezza] = costo

                        self.barre_disponibili.append((qty, lunghezza))
                        costo_display = f"{costo:.2f}" if costo is not None and costo != "" else "-"
                        batch_data.append((qty, lunghezza, costo_display))
                        righe_importate += 1

                    else:
                        lunghezza = row[0]
                        costo = row[1] if len(row) > 1 else None

                        if lunghezza is None:
                            continue

                        lunghezza = float(lunghezza)

                        if lunghezza <= 0:
                            if len(errori) < 100:
                                errori.append(f"Riga {idx}: lunghezza deve essere positiva")
                            continue

                        if lunghezza in self.lunghezze_catalogo:
                            if len(errori) < 100:
                                errori.append(f"Riga {idx}: lunghezza {lunghezza} già presente")
                            continue

                        if costo is not None and costo != "":
                            costo = float(costo)
                            if costo >= 0:
                                self.costi_barre[lunghezza] = costo

                        self.lunghezze_catalogo.append(lunghezza)
                        costo_display = f"{costo:.2f}" if costo is not None and costo != "" else "-"
                        batch_data.append(("Catalogo", lunghezza, costo_display))
                        righe_importate += 1

                    # Inserimento batch per performance
                    if len(batch_data) >= batch_size:
                        for data in batch_data:
                            self.tree_barre.insert("", "end", values=data)
                        batch_data.clear()
                        self.root.update_idletasks()  # Aggiorna GUI periodicamente

                except (ValueError, TypeError) as e:
                    if len(errori) < 100:
                        errori.append(f"Riga {idx}: errore nei dati - {str(e)}")
                    continue

            # Inserisci i dati rimanenti
            for data in batch_data:
                self.tree_barre.insert("", "end", values=data)

            wb.close()

            # Mostra risultati
            msg = f"Importazione completata!\n\nRighe importate: {righe_importate}"
            if errori:
                num_errori = min(len(errori), 100)
                msg += f"\n\nErrori trovati (primi {num_errori}):\n" + "\n".join(errori[:5])
                if len(errori) > 5:
                    msg += f"\n... e altri {len(errori) - 5} errori"
                messagebox.showwarning("Importazione completata con errori", msg)
            else:
                messagebox.showinfo("Successo", msg)

        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante l'importazione del file Excel:\n{str(e)}")

    def importa_pezzi_excel(self):
        """Importa pezzi richiesti da file Excel (ottimizzato)"""
        filename = filedialog.askopenfilename(
            title="Seleziona file Excel con pezzi da tagliare",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not filename:
            return

        try:
            wb = load_workbook(filename, data_only=True, read_only=True)

            # Cerca il foglio giusto
            ws = None
            for nome_foglio in ["Pezzi", "Tagli", "Pezzi Richiesti", "Lista Tagli"]:
                if nome_foglio in wb.sheetnames:
                    ws = wb[nome_foglio]
                    break

            if ws is None:
                ws = wb.active

            righe_importate = 0
            errori = []

            # Batch insert per performance
            batch_data = []
            batch_size = 50

            # Salta l'intestazione e leggi i dati (Quantità, Lunghezza)
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue

                try:
                    qty = row[0]
                    lunghezza = row[1]

                    if qty is None or lunghezza is None:
                        continue

                    qty = int(float(qty))
                    lunghezza = float(lunghezza)

                    if qty <= 0 or lunghezza <= 0:
                        if len(errori) < 100:
                            errori.append(f"Riga {idx}: valori devono essere positivi")
                        continue

                    self.pezzi_richiesti.append((qty, lunghezza))
                    batch_data.append((qty, lunghezza))
                    righe_importate += 1

                    # Inserimento batch
                    if len(batch_data) >= batch_size:
                        for data in batch_data:
                            self.tree_pezzi.insert("", "end", values=data)
                        batch_data.clear()
                        self.root.update_idletasks()

                except (ValueError, TypeError) as e:
                    if len(errori) < 100:
                        errori.append(f"Riga {idx}: errore nei dati - {str(e)}")
                    continue

            # Inserisci i dati rimanenti
            for data in batch_data:
                self.tree_pezzi.insert("", "end", values=data)

            wb.close()

            # Mostra risultati
            msg = f"Importazione completata!\n\nRighe importate: {righe_importate}"
            if errori:
                msg += f"\n\nErrori trovati ({len(errori)}):\n" + "\n".join(errori[:5])
                if len(errori) > 5:
                    msg += f"\n... e altri {len(errori) - 5} errori"
                messagebox.showwarning("Importazione completata con errori", msg)
            else:
                messagebox.showinfo("Successo", msg)

        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante l'importazione del file Excel:\n{str(e)}")

    def crea_excel_esempio(self):
        """Crea file Excel di esempio per l'importazione"""
        # Chiedi dove salvare i file
        directory = filedialog.askdirectory(title="Seleziona cartella dove salvare i file Excel di esempio")

        if not directory:
            return

        try:
            # === FILE 1: Barre Disponibili (Magazzino) ===
            wb1 = Workbook()
            ws1 = wb1.active
            ws1.title = "Barre"

            # Intestazioni
            ws1['A1'] = "Quantità"
            ws1['B1'] = "Lunghezza (mm)"
            ws1['C1'] = "Costo (€)"

            # Stile intestazioni
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ['A1', 'B1', 'C1']:
                ws1[cell].fill = header_fill
                ws1[cell].font = header_font
                ws1[cell].alignment = Alignment(horizontal='center', vertical='center')

            # Dati di esempio
            dati_magazzino = [
                [5, 6000, 28.50],
                [3, 12000, 52.00],
                [10, 3000, 15.00],
                [2, 9000, 42.00]
            ]

            for idx, row_data in enumerate(dati_magazzino, start=2):
                ws1[f'A{idx}'] = row_data[0]
                ws1[f'B{idx}'] = row_data[1]
                ws1[f'C{idx}'] = row_data[2]

            # Larghezza colonne
            ws1.column_dimensions['A'].width = 12
            ws1.column_dimensions['B'].width = 18
            ws1.column_dimensions['C'].width = 15

            filepath1 = os.path.join(directory, "esempio_barre_magazzino.xlsx")
            wb1.save(filepath1)

            # === FILE 2: Catalogo Venditore ===
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.title = "Catalogo"

            # Intestazioni
            ws2['A1'] = "Lunghezza (mm)"
            ws2['B1'] = "Costo (€)"

            # Stile intestazioni
            for cell in ['A1', 'B1']:
                ws2[cell].fill = header_fill
                ws2[cell].font = header_font
                ws2[cell].alignment = Alignment(horizontal='center', vertical='center')

            # Dati di esempio
            dati_catalogo = [
                [3000, 15.00],
                [6000, 28.50],
                [9000, 42.00],
                [12000, 52.00]
            ]

            for idx, row_data in enumerate(dati_catalogo, start=2):
                ws2[f'A{idx}'] = row_data[0]
                ws2[f'B{idx}'] = row_data[1]

            # Larghezza colonne
            ws2.column_dimensions['A'].width = 18
            ws2.column_dimensions['B'].width = 15

            filepath2 = os.path.join(directory, "esempio_catalogo_venditore.xlsx")
            wb2.save(filepath2)

            # === FILE 3: Pezzi Richiesti (Lista Tagli) ===
            wb3 = Workbook()
            ws3 = wb3.active
            ws3.title = "Pezzi"

            # Intestazioni
            ws3['A1'] = "Quantità"
            ws3['B1'] = "Lunghezza (mm)"

            # Stile intestazioni
            for cell in ['A1', 'B1']:
                ws3[cell].fill = header_fill
                ws3[cell].font = header_font
                ws3[cell].alignment = Alignment(horizontal='center', vertical='center')

            # Dati di esempio
            dati_pezzi = [
                [4, 2100],
                [8, 900],
                [2, 5500],
                [6, 1200],
                [3, 3400]
            ]

            for idx, row_data in enumerate(dati_pezzi, start=2):
                ws3[f'A{idx}'] = row_data[0]
                ws3[f'B{idx}'] = row_data[1]

            # Larghezza colonne
            ws3.column_dimensions['A'].width = 12
            ws3.column_dimensions['B'].width = 18

            filepath3 = os.path.join(directory, "esempio_pezzi_richiesti.xlsx")
            wb3.save(filepath3)

            messagebox.showinfo("Successo",
                              f"File Excel di esempio creati con successo!\n\n"
                              f"1. {os.path.basename(filepath1)}\n"
                              f"   → Usalo in modalità 'Barre disponibili'\n\n"
                              f"2. {os.path.basename(filepath2)}\n"
                              f"   → Usalo in modalità 'Calcola fabbisogno'\n\n"
                              f"3. {os.path.basename(filepath3)}\n"
                              f"   → Per importare i pezzi da tagliare\n\n"
                              f"Percorso: {directory}")

        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante la creazione dei file Excel:\n{str(e)}")

    def genera_tutti_scenari(self, pezzi_richiesti, lunghezze_catalogo, spessore_lama, costi_barre=None):
        """
        Genera tutti gli scenari possibili di taglio provando diverse combinazioni di barre.

        Args:
            pezzi_richiesti: Lista di tuple (quantità, lunghezza)
            lunghezze_catalogo: Lista delle lunghezze disponibili nel catalogo
            spessore_lama: Spessore della lama in mm
            costi_barre: Dict opzionale {lunghezza: costo} per calcolare il costo totale

        Returns:
            Lista di scenari ordinati per spreco crescente. Ogni scenario contiene:
            - fabbisogno: dict {lunghezza: quantità}
            - spreco_totale: float
            - scarti: lista delle lunghezze degli scarti
            - num_barre_totale: int
            - costo_totale: float (solo se costi_barre è fornito)
            - barre_dettaglio: lista delle barre con tagli
        """
        import itertools

        # Espandi i pezzi
        tutti_pezzi = []
        for qty, lunghezza in pezzi_richiesti:
            tutti_pezzi.extend([lunghezza] * qty)

        # Ordina per lunghezza decrescente
        tutti_pezzi.sort(reverse=True)

        scenari = []

        # Strategia 1: Scenario con spreco minimo (algoritmo greedy esistente)
        scenario_minimo = self._calcola_scenario_greedy(tutti_pezzi, lunghezze_catalogo, spessore_lama, costi_barre)
        scenari.append(scenario_minimo)

        # Strategia 2: Prova diverse combinazioni forzando l'uso di barre diverse
        # Per ogni lunghezza di barra, prova a creare scenari che privilegiano quella lunghezza
        for lung_preferita in sorted(lunghezze_catalogo, reverse=True):
            scenario = self._calcola_scenario_con_preferenza(tutti_pezzi, lunghezze_catalogo,
                                                             spessore_lama, lung_preferita, costi_barre)
            if scenario and not self._scenario_duplicato(scenario, scenari):
                scenari.append(scenario)

        # Strategia 3: Scenario con numero minimo di barre (privilegia barre lunghe)
        scenario_min_barre = self._calcola_scenario_min_barre(tutti_pezzi, lunghezze_catalogo,
                                                               spessore_lama, costi_barre)
        if scenario_min_barre and not self._scenario_duplicato(scenario_min_barre, scenari):
            scenari.append(scenario_min_barre)

        # Strategia 4: Scenario con scarti più lunghi
        scenario_scarti_lunghi = self._calcola_scenario_scarti_lunghi(tutti_pezzi, lunghezze_catalogo,
                                                                       spessore_lama, costi_barre)
        if scenario_scarti_lunghi and not self._scenario_duplicato(scenario_scarti_lunghi, scenari):
            scenari.append(scenario_scarti_lunghi)

        # Ordina gli scenari per spreco crescente
        scenari.sort(key=lambda x: x['spreco_totale'])

        return scenari

    def _scenario_duplicato(self, scenario, lista_scenari):
        """Verifica se uno scenario è duplicato (stesso fabbisogno)"""
        for s in lista_scenari:
            if s['fabbisogno'] == scenario['fabbisogno']:
                return True
        return False

    def _calcola_scenario_greedy(self, tutti_pezzi, lunghezze_catalogo, spessore_lama, costi_barre):
        """Algoritmo greedy: minimizza lo spreco per singola barra"""
        lunghezze_ord = sorted(lunghezze_catalogo)
        barre_utilizzate = []
        pezzi_usati = set()
        n_pezzi = len(tutti_pezzi)

        while len(pezzi_usati) < n_pezzi:
            # Trova primo pezzo non usato (ottimizzato)
            pezzo_idx = next(i for i in range(n_pezzi) if i not in pezzi_usati)
            pezzo = tutti_pezzi[pezzo_idx]
            inserito = False

            # Cerca barra già aperta (early exit se trovata)
            miglior_barra_idx = -1
            miglior_spreco_barra = float('inf')

            for idx, barra in enumerate(barre_utilizzate):
                if pezzo <= barra['spazio_rimanente']:
                    spreco = barra['spazio_rimanente'] - pezzo - spessore_lama
                    if spreco < miglior_spreco_barra:
                        miglior_spreco_barra = spreco
                        miglior_barra_idx = idx
                        if spreco == 0:  # Perfect fit, ferma la ricerca
                            break

            if miglior_barra_idx >= 0:
                barra_scelta = barre_utilizzate[miglior_barra_idx]
                barra_scelta['pezzi'].append(pezzo)
                barra_scelta['spazio_rimanente'] -= (pezzo + spessore_lama)
                pezzi_usati.add(pezzo_idx)
                inserito = True

            # Apri nuova barra
            if not inserito:
                barre_compatibili = [lung for lung in lunghezze_ord if pezzo <= lung]
                if not barre_compatibili:
                    raise ValueError(f"Nessuna barra può contenere il pezzo da {pezzo}mm")

                migliore_barra = barre_compatibili[0]
                miglior_spreco = float('inf')
                migliori_pezzi = [pezzo]
                migliori_idx = [pezzo_idx]

                # Ottimizzazione: limita il numero di barre da testare per grandi dataset
                max_test = min(3, len(barre_compatibili))

                for lung_barra in barre_compatibili[:max_test]:
                    spazio_sim = lung_barra - pezzo - spessore_lama
                    pezzi_sim = [pezzo]
                    idx_sim = [pezzo_idx]

                    # Ottimizzazione: limita quanti pezzi aggiuntivi cercare
                    pezzi_aggiunti = 0
                    max_pezzi_extra = 10

                    for i in range(n_pezzi):
                        if pezzi_aggiunti >= max_pezzi_extra:
                            break
                        if i not in pezzi_usati and i != pezzo_idx:
                            p = tutti_pezzi[i]
                            if p <= spazio_sim:
                                pezzi_sim.append(p)
                                idx_sim.append(i)
                                spazio_sim -= (p + spessore_lama)
                                pezzi_aggiunti += 1

                    if spazio_sim < miglior_spreco:
                        miglior_spreco = spazio_sim
                        migliore_barra = lung_barra
                        migliori_pezzi = pezzi_sim
                        migliori_idx = idx_sim

                nuova_barra = {
                    'lunghezza': migliore_barra,
                    'pezzi': migliori_pezzi,
                    'spazio_rimanente': miglior_spreco
                }
                barre_utilizzate.append(nuova_barra)
                pezzi_usati.update(migliori_idx)

        return self._crea_scenario(barre_utilizzate, lunghezze_catalogo, costi_barre)

    def _calcola_scenario_con_preferenza(self, tutti_pezzi, lunghezze_catalogo, spessore_lama,
                                        lung_preferita, costi_barre):
        """Calcola scenario privilegiando una specifica lunghezza di barra"""
        lunghezze_ord = [lung_preferita] + [l for l in sorted(lunghezze_catalogo) if l != lung_preferita]
        barre_utilizzate = []
        pezzi_usati = set()
        n_pezzi = len(tutti_pezzi)

        while len(pezzi_usati) < n_pezzi:
            pezzo_idx = next(i for i in range(n_pezzi) if i not in pezzi_usati)
            pezzo = tutti_pezzi[pezzo_idx]
            inserito = False

            # Cerca barra già aperta
            barre_compatibili = [(idx, barra) for idx, barra in enumerate(barre_utilizzate)
                                if pezzo <= barra['spazio_rimanente']]

            if barre_compatibili:
                # Privilegia barre della lunghezza preferita
                barre_preferite = [(idx, b) for idx, b in barre_compatibili if b['lunghezza'] == lung_preferita]
                if barre_preferite:
                    barre_compatibili = barre_preferite

                barre_compatibili.sort(key=lambda x: x[1]['spazio_rimanente'] - pezzo - spessore_lama)
                idx_scelto, barra_scelta = barre_compatibili[0]
                barra_scelta['pezzi'].append(pezzo)
                barra_scelta['spazio_rimanente'] -= (pezzo + spessore_lama)
                pezzi_usati.add(pezzo_idx)
                inserito = True

            if not inserito:
                # Usa la lunghezza preferita se possibile
                barre_compatibili = [lung for lung in lunghezze_ord if pezzo <= lung]
                if not barre_compatibili:
                    return None

                lung_barra = barre_compatibili[0]
                spazio_sim = lung_barra - pezzo - spessore_lama
                pezzi_sim = [pezzo]
                idx_sim = [pezzo_idx]

                for i in range(n_pezzi):
                    if i not in pezzi_usati and i != pezzo_idx:
                        p = tutti_pezzi[i]
                        if p <= spazio_sim:
                            pezzi_sim.append(p)
                            idx_sim.append(i)
                            spazio_sim -= (p + spessore_lama)

                nuova_barra = {
                    'lunghezza': lung_barra,
                    'pezzi': pezzi_sim,
                    'spazio_rimanente': spazio_sim
                }
                barre_utilizzate.append(nuova_barra)
                pezzi_usati.update(idx_sim)

        return self._crea_scenario(barre_utilizzate, lunghezze_catalogo, costi_barre)

    def _calcola_scenario_min_barre(self, tutti_pezzi, lunghezze_catalogo, spessore_lama, costi_barre):
        """Calcola scenario che minimizza il numero di barre (usa barre più lunghe)"""
        lunghezze_ord = sorted(lunghezze_catalogo, reverse=True)  # Privilegia barre lunghe
        barre_utilizzate = []
        pezzi_usati = set()
        n_pezzi = len(tutti_pezzi)

        while len(pezzi_usati) < n_pezzi:
            pezzo_idx = next(i for i in range(n_pezzi) if i not in pezzi_usati)
            pezzo = tutti_pezzi[pezzo_idx]
            inserito = False

            # Cerca barra già aperta
            barre_compatibili = [(idx, barra) for idx, barra in enumerate(barre_utilizzate)
                                if pezzo <= barra['spazio_rimanente']]

            if barre_compatibili:
                # Privilegia barre più lunghe con più spazio
                barre_compatibili.sort(key=lambda x: (-x[1]['lunghezza'], x[1]['spazio_rimanente'] - pezzo))
                idx_scelto, barra_scelta = barre_compatibili[0]
                barra_scelta['pezzi'].append(pezzo)
                barra_scelta['spazio_rimanente'] -= (pezzo + spessore_lama)
                pezzi_usati.add(pezzo_idx)
                inserito = True

            if not inserito:
                # Usa sempre la barra più lunga possibile
                barre_compatibili = [lung for lung in lunghezze_ord if pezzo <= lung]
                if not barre_compatibili:
                    return None

                lung_barra = barre_compatibili[0]  # La più lunga
                spazio_sim = lung_barra - pezzo - spessore_lama
                pezzi_sim = [pezzo]
                idx_sim = [pezzo_idx]

                for i in range(n_pezzi):
                    if i not in pezzi_usati and i != pezzo_idx:
                        p = tutti_pezzi[i]
                        if p <= spazio_sim:
                            pezzi_sim.append(p)
                            idx_sim.append(i)
                            spazio_sim -= (p + spessore_lama)

                nuova_barra = {
                    'lunghezza': lung_barra,
                    'pezzi': pezzi_sim,
                    'spazio_rimanente': spazio_sim
                }
                barre_utilizzate.append(nuova_barra)
                pezzi_usati.update(idx_sim)

        return self._crea_scenario(barre_utilizzate, lunghezze_catalogo, costi_barre)

    def _calcola_scenario_scarti_lunghi(self, tutti_pezzi, lunghezze_catalogo, spessore_lama, costi_barre):
        """Calcola scenario che privilegia scarti più lunghi e riutilizzabili"""
        lunghezze_ord = sorted(lunghezze_catalogo)
        barre_utilizzate = []
        pezzi_usati = set()
        n_pezzi = len(tutti_pezzi)

        while len(pezzi_usati) < n_pezzi:
            pezzo_idx = next(i for i in range(n_pezzi) if i not in pezzi_usati)
            pezzo = tutti_pezzi[pezzo_idx]
            inserito = False

            # Cerca barra già aperta - ma evita di riempirle completamente se possibile
            barre_compatibili = [(idx, barra) for idx, barra in enumerate(barre_utilizzate)
                                if pezzo <= barra['spazio_rimanente']]

            if barre_compatibili:
                # Preferisci barre che dopo il taglio lasciano scarti > 500mm oppure < 100mm
                # (o molto riutilizzabili o quasi zero)
                def priorita_scarto(x):
                    scarto_dopo = x[1]['spazio_rimanente'] - pezzo - spessore_lama
                    if scarto_dopo > 500:
                        return 0  # Ottimo, scarto riutilizzabile
                    elif scarto_dopo < 100:
                        return 1  # Buono, quasi zero spreco
                    else:
                        return 2  # Peggiore, scarto medio inutilizzabile

                barre_compatibili.sort(key=priorita_scarto)
                idx_scelto, barra_scelta = barre_compatibili[0]
                barra_scelta['pezzi'].append(pezzo)
                barra_scelta['spazio_rimanente'] -= (pezzo + spessore_lama)
                pezzi_usati.add(pezzo_idx)
                inserito = True

            if not inserito:
                # Scegli barra che massimizza lo scarto finale riutilizzabile
                barre_compatibili = [lung for lung in lunghezze_ord if pezzo <= lung]
                if not barre_compatibili:
                    return None

                migliore_barra = barre_compatibili[0]  # Default: prima barra compatibile
                miglior_score = -float('inf')  # Score iniziale molto basso
                migliori_pezzi = [pezzo]  # Default: solo il pezzo corrente
                migliori_idx = [pezzo_idx]

                for lung_barra in barre_compatibili:
                    spazio_sim = lung_barra - pezzo - spessore_lama
                    pezzi_sim = [pezzo]
                    idx_sim = [pezzo_idx]

                    for i in range(n_pezzi):
                        if i not in pezzi_usati and i != pezzo_idx:
                            p = tutti_pezzi[i]
                            if p <= spazio_sim:
                                pezzi_sim.append(p)
                                idx_sim.append(i)
                                spazio_sim -= (p + spessore_lama)

                    # Score: privilegia scarti > 500mm
                    if spazio_sim > 500:
                        score = spazio_sim  # Più lungo è meglio
                    else:
                        score = -spazio_sim  # Più corto è meglio

                    if score > miglior_score:
                        miglior_score = score
                        migliore_barra = lung_barra
                        migliori_pezzi = pezzi_sim.copy()
                        migliori_idx = idx_sim.copy()

                nuova_barra = {
                    'lunghezza': migliore_barra,
                    'pezzi': migliori_pezzi,
                    'spazio_rimanente': migliore_barra - sum(migliori_pezzi) - len(migliori_pezzi) * spessore_lama
                }
                barre_utilizzate.append(nuova_barra)
                pezzi_usati.update(migliori_idx)

        return self._crea_scenario(barre_utilizzate, lunghezze_catalogo, costi_barre)

    def _crea_scenario(self, barre_utilizzate, lunghezze_catalogo, costi_barre):
        """Crea un oggetto scenario dai dati delle barre utilizzate"""
        # Conta fabbisogno
        fabbisogno = {lung: 0 for lung in lunghezze_catalogo}
        for barra in barre_utilizzate:
            fabbisogno[barra['lunghezza']] += 1

        # Calcola spreco totale e scarti
        spreco_totale = sum(b['spazio_rimanente'] for b in barre_utilizzate)
        scarti = sorted([b['spazio_rimanente'] for b in barre_utilizzate], reverse=True)

        # Calcola costo se fornito
        costo_totale = None
        if costi_barre:
            costo_totale = sum(costi_barre.get(lung, 0) * qty for lung, qty in fabbisogno.items())

        return {
            'fabbisogno': fabbisogno,
            'spreco_totale': spreco_totale,
            'scarti': scarti,
            'num_barre_totale': len(barre_utilizzate),
            'costo_totale': costo_totale,
            'barre_dettaglio': barre_utilizzate
        }

    def ottimizza(self):
        if not self.pezzi_richiesti:
            messagebox.showwarning("Attenzione", "Inserire almeno un pezzo da tagliare")
            return

        modalita = self.modalita.get()

        if modalita == "disponibili":
            if not self.barre_disponibili:
                messagebox.showwarning("Attenzione", "Inserire almeno una barra disponibile")
                return
            # Modalità disponibili: usa algoritmo classico
            self._ottimizza_con_barre_disponibili()
        else:
            if not self.lunghezze_catalogo:
                messagebox.showwarning("Attenzione", "Inserire almeno una lunghezza nel catalogo")
                return
            # Modalità calcola: mostra tutti gli scenari
            self._mostra_scenari()

    def _ottimizza_con_barre_disponibili(self):
        """Ottimizzazione classica con barre già disponibili"""
        try:
            spessore_lama = float(self.entry_spessore_lama.get())

            # Validazione: verifica che tutte le barre possano contenere almeno un pezzo
            max_lunghezza_barra = max(lung for _, lung in self.barre_disponibili)
            max_pezzo = max(lung for _, lung in self.pezzi_richiesti)

            if max_pezzo > max_lunghezza_barra:
                messagebox.showerror("Errore",
                    f"Errore: c'è un pezzo da {max_pezzo}mm che è più lungo della barra più lunga ({max_lunghezza_barra}mm)!\n\n"
                    f"Soluzione: aggiungi barre più lunghe di almeno {max_pezzo}mm")
                return

            # Esegui ottimizzazione
            ottimizzatore = OttimizzatoreTaglio(self.barre_disponibili, spessore_lama)
            barre = ottimizzatore.ottimizza(self.pezzi_richiesti)

            # Salva risultati
            self.risultati_ottimizzazione = barre

            # Pulisci risultati precedenti
            self.tree_risultati.delete(*self.tree_risultati.get_children())

            # Mostra risultati
            tot_sfrido = 0
            tot_pezzi = len(barre)
            lunghezza_totale = 0
            costo_barre_intere = 0
            costo_effettivo = 0

            for i, barra in enumerate(barre, 1):
                tagli_str = " + ".join(str(int(t)) for t in barra['tagli'])
                self.tree_risultati.insert("", "end", values=(
                    i,
                    f"{int(barra['lunghezza'])} mm",
                    tagli_str,
                    barra['num_tagli'],
                    f"{barra['sfrido']:.1f}"
                ))
                tot_sfrido += barra['sfrido']
                lunghezza_totale += barra['lunghezza']

                # Calcola costi se disponibili
                lung_barra = barra['lunghezza']
                if lung_barra in self.costi_barre:
                    costo_barra = self.costi_barre[lung_barra]
                    costo_barre_intere += costo_barra

                    # Costo effettivo = (costo/lunghezza) * lunghezza_utilizzata
                    # lunghezza_utilizzata = lunghezza_barra - sfrido
                    costo_per_mm = costo_barra / lung_barra
                    lunghezza_utilizzata = lung_barra - barra['sfrido']
                    costo_effettivo += costo_per_mm * lunghezza_utilizzata

            # Salva i costi per il PDF
            self.costo_barre_intere = costo_barre_intere
            self.costo_effettivo = costo_effettivo

            # Mostra statistiche
            efficienza = ((lunghezza_totale - tot_sfrido) / lunghezza_totale * 100) if lunghezza_totale > 0 else 0
            stats_text = f"Barre utilizzate: {tot_pezzi} | Sfrido totale: {tot_sfrido:.1f} mm | Efficienza: {efficienza:.1f}%"

            # Aggiungi costi se disponibili
            if costo_barre_intere > 0:
                stats_text += f" | Costo barre intere: €{costo_barre_intere:.2f} | Costo effettivo: €{costo_effettivo:.2f}"

            self.label_stats.config(text=stats_text)

        except ValueError as e:
            messagebox.showerror("Errore", f"Errore durante l'ottimizzazione:\n{str(e)}")
        except Exception as e:
            messagebox.showerror("Errore", f"Errore imprevisto:\n{str(e)}\n\nDettagli tecnici: {type(e).__name__}")

    def _mostra_scenari(self):
        """Mostra finestra con tutti gli scenari possibili"""
        try:
            spessore_lama = float(self.entry_spessore_lama.get())

            # Genera tutti gli scenari
            costi_opzionali = self.costi_barre if self.costi_barre else None
            scenari = self.genera_tutti_scenari(self.pezzi_richiesti, self.lunghezze_catalogo,
                                               spessore_lama, costi_opzionali)

            # Crea finestra modale per scenari
            self._crea_finestra_scenari(scenari, spessore_lama)

        except ValueError as e:
            messagebox.showerror("Errore", f"Errore durante il calcolo degli scenari:\n{str(e)}")
        except Exception as e:
            messagebox.showerror("Errore", f"Errore imprevisto durante il calcolo degli scenari:\n{str(e)}\n\nDettagli: {type(e).__name__}")

    def _crea_finestra_scenari(self, scenari, spessore_lama):
        """Crea finestra modale per mostrare tutti gli scenari"""
        finestra_scenari = tk.Toplevel(self.root)
        finestra_scenari.title("Scenari di Taglio - Seleziona il migliore")
        finestra_scenari.geometry("900x700")
        finestra_scenari.transient(self.root)
        finestra_scenari.grab_set()

        # Imposta la stessa icona della finestra principale
        self.imposta_icona(finestra_scenari)

        # Titolo
        titolo_frame = ttk.Frame(finestra_scenari, padding=10)
        titolo_frame.pack(fill="x")
        ttk.Label(titolo_frame, text="COMBINAZIONI POSSIBILI",
                 font=("Arial", 12, "bold")).pack()
        

        # Frame scrollabile per scenari
        canvas = tk.Canvas(finestra_scenari)
        scrollbar = ttk.Scrollbar(finestra_scenari, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Crea card per ogni scenario
        for idx, scenario in enumerate(scenari, 1):
            self._crea_card_scenario(scrollable_frame, idx, scenario, spessore_lama)

        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)

        # Bottone chiudi
        ttk.Button(finestra_scenari, text="Chiudi", command=finestra_scenari.destroy,
                  width=20).pack(pady=10)

    def _crea_card_scenario(self, parent, numero, scenario, spessore_lama):
        """Crea una card per un singolo scenario"""
        # Frame principale con bordo
        card = ttk.LabelFrame(parent, text=f"COMBINAZIONE {numero}", padding=10)
        card.pack(fill="x", padx=5, pady=5)

        # Prima riga: Fabbisogno barre
        fabb_frame = ttk.Frame(card)
        fabb_frame.pack(fill="x", pady=(0, 5))

        fabb_text = "Barre da ordinare: "
        fabb_items = []
        for lung, qty in sorted(scenario['fabbisogno'].items(), reverse=True):
            if qty > 0:
                fabb_items.append(f"{qty}×{int(lung)}mm")
        fabb_text += ", ".join(fabb_items)

        ttk.Label(fabb_frame, text=fabb_text, font=("Arial", 10, "bold")).pack(anchor="w")

        # Seconda riga: Statistiche
        stats_frame = ttk.Frame(card)
        stats_frame.pack(fill="x", pady=2)

        stats_left = ttk.Frame(stats_frame)
        stats_left.pack(side="left", fill="x", expand=True)

        ttk.Label(stats_left, text=f"• Numero barre: {scenario['num_barre_totale']}").pack(anchor="w")
        ttk.Label(stats_left, text=f"• Spreco totale: {scenario['spreco_totale']:.1f} mm").pack(anchor="w")

        # Scarti
        scarti_str = ", ".join([f"{s:.0f}mm" for s in scenario['scarti'][:5]])
        if len(scenario['scarti']) > 5:
            scarti_str += f" (+{len(scenario['scarti'])-5} altri)"
        ttk.Label(stats_left, text=f"• Scarti: {scarti_str}").pack(anchor="w")

        # Costo se disponibile
        if scenario['costo_totale'] is not None:
            ttk.Label(stats_left, text=f"• Costo totale: €{scenario['costo_totale']:.2f}",
                     font=("Arial", 9, "bold"), foreground="green").pack(anchor="w")

        # Bottone per selezionare questo scenario
        btn_frame = ttk.Frame(card)
        btn_frame.pack(fill="x", pady=(5, 0))

        ttk.Button(btn_frame, text="SELEZIONA",
                  command=lambda s=scenario: self._seleziona_scenario(s, spessore_lama)).pack()

    def _seleziona_scenario(self, scenario, spessore_lama):
        """Seleziona uno scenario e mostra i dettagli di taglio"""
        # Converti scenario in formato barre_disponibili
        barre_disponibili = [(qty, lung) for lung, qty in scenario['fabbisogno'].items() if qty > 0]

        # Esegui ottimizzazione con queste barre
        ottimizzatore = OttimizzatoreTaglio(barre_disponibili, spessore_lama)
        barre = ottimizzatore.ottimizza(self.pezzi_richiesti)

        # Salva risultati
        self.risultati_ottimizzazione = barre

        # Pulisci risultati precedenti
        self.tree_risultati.delete(*self.tree_risultati.get_children())

        # Mostra risultati
        tot_sfrido = 0
        tot_pezzi = len(barre)
        lunghezza_totale = 0

        for i, barra in enumerate(barre, 1):
            tagli_str = " + ".join(str(int(t)) for t in barra['tagli'])
            self.tree_risultati.insert("", "end", values=(
                i,
                f"{int(barra['lunghezza'])} mm",
                tagli_str,
                barra['num_tagli'],
                f"{barra['sfrido']:.1f}"
            ))
            tot_sfrido += barra['sfrido']
            lunghezza_totale += barra['lunghezza']

        # Mostra statistiche
        efficienza = ((lunghezza_totale - tot_sfrido) / lunghezza_totale * 100) if lunghezza_totale > 0 else 0

        # Aggiungi info sul fabbisogno e costo
        fabbisogno_str = " | Fabbisogno: "
        # Conta lunghezze in un solo passaggio usando dict
        fabb_count = {}
        for b in barre:
            lung = b['lunghezza']
            fabb_count[lung] = fabb_count.get(lung, 0) + 1

        fabb_items = [f"{count}x{int(lung)}mm" for lung, count in sorted(fabb_count.items(), reverse=True)]

        stats_text = f"Barre utilizzate: {tot_pezzi} | Sfrido totale: {tot_sfrido:.1f} mm | Efficienza: {efficienza:.1f}%"
        stats_text += fabbisogno_str + ", ".join(fabb_items)

        if scenario['costo_totale'] is not None:
            stats_text += f" | Costo: €{scenario['costo_totale']:.2f}"

        self.label_stats.config(text=stats_text)

        # Chiudi tutte le finestre TopLevel
        for widget in self.root.winfo_children():
            if isinstance(widget, tk.Toplevel):
                widget.destroy()

        messagebox.showinfo("Scenario Selezionato",
                           "Scenario applicato con successo!\nDettagli visibili nella tabella sottostante.")

    def genera_pdf(self):
        if not self.risultati_ottimizzazione:
            messagebox.showwarning("Attenzione", "Eseguire prima l'ottimizzazione")
            return

        # Mostra finestra per inserire nome progetto e data
        dialogo = tk.Toplevel(self.root)
        dialogo.title("Informazioni PDF")
        dialogo.geometry("400x200")
        dialogo.transient(self.root)
        dialogo.grab_set()

        # Imposta icona
        self.imposta_icona(dialogo)

        # Frame contenuto
        frame_content = ttk.Frame(dialogo, padding=20)
        frame_content.pack(fill="both", expand=True)

        # Nome progetto
        ttk.Label(frame_content, text="Nome Progetto:", font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=5)
        entry_progetto = ttk.Entry(frame_content, width=30)
        entry_progetto.insert(0, "Piano di Taglio")
        entry_progetto.grid(row=0, column=1, pady=5, padx=5)

        # Data
        ttk.Label(frame_content, text="Data:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=5)
        entry_data = ttk.Entry(frame_content, width=30)
        entry_data.insert(0, datetime.now().strftime('%d/%m/%Y'))
        entry_data.grid(row=1, column=1, pady=5, padx=5)

        # Variabile per salvare i dati
        dati_pdf = {'confermato': False, 'progetto': '', 'data': ''}

        def conferma():
            dati_pdf['confermato'] = True
            dati_pdf['progetto'] = entry_progetto.get().strip()
            dati_pdf['data'] = entry_data.get().strip()
            dialogo.destroy()

        def annulla():
            dialogo.destroy()

        # Frame pulsanti
        frame_buttons = ttk.Frame(frame_content)
        frame_buttons.grid(row=2, column=0, columnspan=2, pady=20)

        ttk.Button(frame_buttons, text="Conferma", command=conferma, width=15).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="Annulla", command=annulla, width=15).pack(side="left", padx=5)

        # Centra la finestra
        dialogo.wait_window()

        # Se l'utente ha annullato
        if not dati_pdf['confermato']:
            return

        # Usa valori di default se vuoti
        nome_progetto = dati_pdf['progetto'] if dati_pdf['progetto'] else "Piano di Taglio"
        data_progetto = dati_pdf['data'] if dati_pdf['data'] else datetime.now().strftime('%d/%m/%Y')

        # Chiedi dove salvare il file
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=f"piano_taglio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if not filename:
            return

        try:
            # Crea il PDF con margini ridotti
            doc = SimpleDocTemplate(
                filename,
                pagesize=A4,
                topMargin=15*mm,
                bottomMargin=15*mm,
                leftMargin=15*mm,
                rightMargin=15*mm
            )
            story = []
            styles = getSampleStyleSheet()

            # Calcola statistiche
            tot_sfrido = sum(barra['sfrido'] for barra in self.risultati_ottimizzazione)
            lunghezza_totale = sum(barra['lunghezza'] for barra in self.risultati_ottimizzazione)
            efficienza = ((lunghezza_totale - tot_sfrido) / lunghezza_totale * 100) if lunghezza_totale > 0 else 0

            # Intestazione compatta
            header_style = ParagraphStyle(
                'Header',
                parent=styles['Heading1'],
                fontSize=14,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=8,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            story.append(Paragraph(nome_progetto.upper(), header_style))

            # Info progetto
            progetto_style = ParagraphStyle('Progetto', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'))
            story.append(Paragraph(f"Data: {data_progetto}", progetto_style))
            story.append(Spacer(1, 5))

            # Info in una riga
            info_text = f"Lama: {self.entry_spessore_lama.get()}mm | Barre: {len(self.risultati_ottimizzazione)} | Sfrido: {tot_sfrido:.0f}mm | Efficienza: {efficienza:.1f}%"
            info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)
            story.append(Paragraph(info_text, info_style))

            # Aggiungi costi se disponibili
            if self.costo_barre_intere > 0:
                costo_text = f"Costo barre intere: €{self.costo_barre_intere:.2f} | Costo effettivo (materiale utilizzato): €{self.costo_effettivo:.2f}"
                costo_style = ParagraphStyle('Costo', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#27ae60'))
                story.append(Paragraph(costo_text, costo_style))

            story.append(Spacer(1, 10))

            # Tabella principale con tutte le barre
            main_data = [["Barra", "Lung.", "Tagli", "Sfrido"]]

            for i, barra in enumerate(self.risultati_ottimizzazione, 1):
                tagli_str = " + ".join(str(int(t)) for t in barra['tagli'])
                main_data.append([
                    f"#{i}",
                    f"{int(barra['lunghezza'])}",
                    tagli_str,
                    f"{barra['sfrido']:.0f}"
                ])

            # Calcola larghezza dinamica per la colonna tagli
            tagli_width = 400
            main_table = Table(main_data, colWidths=[35, 50, tagli_width, 50])
            main_table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Barra centrata
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # Lunghezza centrata
                ('ALIGN', (2, 0), (2, -1), 'LEFT'),    # Tagli allineati a sinistra
                ('ALIGN', (3, 0), (3, -1), 'CENTER'),  # Sfrido centrato
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                # Body
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(main_table)

            # Genera il PDF
            doc.build(story)

            messagebox.showinfo("Successo", f"PDF generato con successo:\n{filename}")

        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante la generazione del PDF:\n{str(e)}")

    def mostra_help(self):
        """Mostra finestra di aiuto con istruzioni per l'uso"""
        help_window = tk.Toplevel(self.root)
        help_window.title("Guida - Ottimizzatore Taglio Barre")
        help_window.geometry("700x600")
        help_window.resizable(False, False)

        # Imposta icona punto interrogativo per la finestra di help
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            help_ico_path = os.path.join(script_dir, "help_icon.ico")
            help_png_path = os.path.join(script_dir, "help_icon.png")

            if os.path.exists(help_ico_path):
                help_window.iconbitmap(help_ico_path)
            elif os.path.exists(help_png_path):
                help_icon = tk.PhotoImage(file=help_png_path)
                help_window.iconphoto(True, help_icon)
        except:
            pass

        # Contenuto guida
        help_text = """
GUIDA ALL'USO - OTTIMIZZATORE TAGLIO BARRE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PANORAMICA
Questo programma ottimizza il taglio di barre,
minimizzando gli scarti e il numero di barre utilizzate.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

IMPORTAZIONE DA EXCEL
✓ Clicca "Crea Excel Esempio" per generare 3 file di esempio:
  1. esempio_barre_magazzino.xlsx
  2. esempio_catalogo_venditore.xlsx
  3. esempio_pezzi_richiesti.xlsx

✓ Modifica i file con i tuoi dati reali
✓ Usa "Importa Excel" per caricare i dati nell'applicazione

IMPORTANTE: Il formato deve rispettare l'intestazione dei file esempio!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PASSO 1: IMPOSTARE I PARAMETRI

• Spessore lama: Inserire lo spessore della lama di taglio in mm
  (es. 3mm per seghe normali, 5mm per seghe più grandi)
  Questo valore verrà sottratto ad ogni taglio.

• Modalità: Scegliere tra due modalità operative:

  → BARRE DISPONIBILI: Usa questa modalità quando hai già
    un numero definito di barre in magazzino e vuoi ottimizzare
    il taglio con quello che hai disponibile.

  → CALCOLA FABBISOGNO: Usa questa modalità quando devi ancora
    acquistare le barre e vuoi sapere QUANTE ne servono.
    Il programma calcolerà automaticamente il numero minimo
    di barre necessarie per ogni lunghezza del catalogo.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PASSO 2A: MODALITÀ "BARRE DISPONIBILI"

Usa questa modalità quando hai barre già disponibili.

INSERIMENTO MANUALE:
1. Inserire la QUANTITÀ di barre disponibili
2. Inserire la LUNGHEZZA della barra in mm
3. Inserire il COSTO (€) - opzionale
4. Cliccare "Aggiungi"

Esempi:
  - 10 barre da 12000mm → Costo 52€
  - 5 barre da 6000mm → Costo 28.50€
  - 3 barre da 3000mm → Costo 15€

IMPORTAZIONE DA EXCEL:
1. Clicca "Importa Excel" nella sezione Barre Disponibili
2. Seleziona il file Excel (es. esempio_barre_magazzino.xlsx)
3. Il file deve avere le colonne:
   Colonna A: Quantità
   Colonna B: Lunghezza (mm)
   Colonna C: Costo (€) - opzionale
4. La prima riga deve contenere le intestazioni
5. I dati iniziano dalla riga 2

Il programma ottimizzerà i tagli usando le barre che hai.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PASSO 2B: MODALITÀ "CALCOLA FABBISOGNO"

Usa questa modalità quando devi acquistare le barre.

INSERIMENTO MANUALE:
1. Inserire solo le LUNGHEZZE disponibili nel catalogo del negozio
2. OPZIONALE: Inserire il COSTO (€) per ogni lunghezza
   (se inserito, vedrai il costo totale di ogni scenario)
3. Cliccare "Aggiungi" per ogni lunghezza
4. NON inserire quantità (il programma le calcolerà!)

Esempi di catalogo:
  - 3000mm → Costo: 15€ (opzionale)
  - 6000mm → Costo: 28€ (opzionale)
  - 12000mm → Costo: 50€ (opzionale)

IMPORTAZIONE DA EXCEL:
1. Clicca "Importa Excel" nella sezione Barre Disponibili
2. Seleziona il file Excel (es. esempio_catalogo_venditore.xlsx)
3. Il file deve avere le colonne:
   Colonna A: Lunghezza (mm)
   Colonna B: Costo (€) - opzionale
4. La prima riga deve contenere le intestazioni
5. I dati iniziano dalla riga 2

Il programma ti mostrerà TUTTI GLI SCENARI POSSIBILI con:
  • Barre da ordinare per ogni scenario
  • Spreco totale
  • Scarti da stoccare (pezzi riutilizzabili)
  • Numero totale di barre
  • Costo totale (se hai inserito i prezzi)

Il PRIMO scenario è sempre quello con SPRECO MINIMO.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PASSO 3: INSERIRE I PEZZI DA TAGLIARE

INSERIMENTO MANUALE:
1. Inserire la QUANTITÀ di pezzi identici richiesti
2. Inserire la LUNGHEZZA del pezzo in mm
3. Cliccare "Aggiungi"

Esempi:
  - 4 pezzi da 2100mm
  - 8 pezzi da 900mm
  - 2 pezzi da 5500mm

IMPORTAZIONE DA EXCEL:
1. Clicca "Importa Excel" nella sezione Pezzi Richiesti
2. Seleziona il file Excel (es. esempio_pezzi_richiesti.xlsx)
3. Il file deve avere le colonne:
   Colonna A: Quantità
   Colonna B: Lunghezza (mm)
4. La prima riga deve contenere le intestazioni
5. I dati iniziano dalla riga 2

Per rimuovere un pezzo: selezionarlo e cliccare "Rimuovi"
Per cancellare tutto: cliccare "Pulisci"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PASSO 4: OTTIMIZZARE

Cliccare il pulsante "OTTIMIZZA" per calcolare il piano di taglio.

• In modalità "BARRE DISPONIBILI":
  Ottimizza i tagli con le barre che hai inserito.
  Ad ogni click viene generato un pattern DIVERSO,
  quindi puoi cliccare più volte per esplorare soluzioni alternative.

• In modalità "CALCOLA FABBISOGNO":
  Il programma ti mostra una FINESTRA CON TUTTI GLI SCENARI POSSIBILI!

  Ogni scenario include:
    → Barre da ordinare (es. "2×6000mm, 1×3000mm")
    → Numero totale di barre
    → Spreco totale
    → Scarti stoccabili (pezzi riutilizzabili)
    → Costo totale (se hai inserito i prezzi)

  Gli scenari sono ordinati dal MINORE al maggiore spreco.
  Il PRIMO scenario è sempre il più efficiente.

  Seleziona lo scenario che preferisci cliccando
  "Usa questo scenario" e il programma genererà
  automaticamente il piano di taglio dettagliato.

Il piano di taglio finale mostra:
  • Numero barra
  • Lunghezza barra utilizzata
  • Sequenza dei tagli da eseguire
  • Numero di tagli per barra
  • Sfrido (scarto) per barra

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PASSO 5: GENERARE IL PDF

Dopo aver ottimizzato, cliccare "GENERA PDF" per creare un
documento stampabile con le istruzioni dettagliate di taglio
da fornire all'operatore.

Il PDF include:
  • Data e parametri di taglio
  • Lista completa dei pezzi richiesti
  • Istruzioni dettagliate per ogni barra
  • Lunghezza rimanente dopo ogni taglio
  • Riepilogo finale con statistiche

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

SUGGERIMENTI

✓ Usa "Calcola fabbisogno" quando devi acquistare le barre
  e vuoi sapere esattamente quante ne servono

✓ Usa "Barre disponibili" quando hai già le barre in magazzino
  e vuoi ottimizzare i tagli

✓ Inserire tutte le lunghezze disponibili nel catalogo per
  ottenere l'ottimizzazione migliore

✓ Cliccare "OTTIMIZZA" più volte per esplorare soluzioni diverse

✓ Verificare che tutti i pezzi siano più corti della barra più lunga

✓ Lo spessore lama dipende dalla sega utilizzata (controllare
  le specifiche tecniche)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ESEMPIO 1 - MODALITÀ "BARRE DISPONIBILI"

Scenario: Hai barre in magazzino

Barre disponibili:
  • 5 x 6000mm
  • 3 x 12000mm

Pezzi richiesti:
  • 4 x 2100mm (montanti)
  • 8 x 900mm (traversi)

Risultato:
  Il programma ottimizza i tagli con le barre che hai,
  risparmiando le più lunghe quando possibile.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ESEMPIO 2 - MODALITÀ "CALCOLA FABBISOGNO"

Scenario: Devi acquistare le barre

Catalogo negozio (lunghezze disponibili):
  • 3000mm
  • 6000mm
  • 12000mm

Pezzi richiesti:
  • 4 x 2100mm (montanti)
  • 8 x 900mm (traversi)

Risultato:
  Il programma ti dice: "Servono 2 barre da 6000mm e 1 da 3000mm"
  Poi genera automaticamente il piano di taglio ottimale.

VANTAGGIO: Sai esattamente cosa ordinare al negozio!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

        # Aggiungi il testo
        text_widget = tk.Text(help_window, wrap="word",
                              font=("Consolas", 9), padx=10, pady=10)
        text_widget.insert("1.0", help_text)
        text_widget.config(state="disabled")  # Rendi il testo read-only
        text_widget.pack(fill="both", expand=True)

        # Centra la finestra
        help_window.transient(self.root)
        help_window.grab_set()


if __name__ == "__main__":
    root = tk.Tk()
    app = ApplicativoGUI(root)
    root.mainloop()
